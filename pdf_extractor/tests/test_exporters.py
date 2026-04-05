from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from lifesaving_pdf_extractor.config import PROVISIONAL_EXPORT_FIELDS
from lifesaving_pdf_extractor.exporters import ExcelExporter, ReviewExporter
from lifesaving_pdf_extractor.models import ExtractedRecord, ExtractionStatus, ReviewFlag
from openpyxl import load_workbook


class ExporterTests(unittest.TestCase):
    def test_csv_and_review_export_write_expected_files(self) -> None:
        record = ExtractedRecord(
            source_file=Path("demo.pdf"),
            page_number=1,
            parser_name="generic",
            status=ExtractionStatus.PARTIAL,
            confidence=0.4,
            data={
                "wettkampfname": "Demo Cup",
                "datum": "2026-04-05",
                "disziplin": "hindernis",
            },
            review_notes=["Parser rules are still provisional."],
            raw_excerpt="Demo text",
        )
        flag = ReviewFlag(code="check_me", message="Needs a human review.", source_file=Path("demo.pdf"), page_number=1)

        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            excel_path = ExcelExporter().export([record], temp_path / "records.xlsx", PROVISIONAL_EXPORT_FIELDS)
            review_path = ReviewExporter().export([record], [flag], temp_path / "review.csv")

            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            excel_rows = list(worksheet.iter_rows(values_only=True))
            review_content = review_path.read_text(encoding="utf-8")

        flattened = " | ".join("" if value is None else str(value) for row in excel_rows for value in row)
        self.assertIn("Demo Cup", flattened)
        self.assertIn("Parser rules are still provisional.", review_content)
        self.assertIn("check_me", review_content)

    def test_csv_export_maps_single_event_into_final_schema(self) -> None:
        record = ExtractedRecord(
            source_file=Path("jrp.pdf"),
            page_number=1,
            parser_name="lifesaving",
            status=ExtractionStatus.COMPLETE,
            confidence=0.9,
            data={
                "wettkampfname": "Junioren Rettungspokal 2025",
                "datum": "2025-06-28",
                "ort": "Bremen",
                "disziplin": "100m Hindernisschwimmen",
                "altersklasse": "AK 17/18",
                "geschlecht": "weiblich",
                "platz": "1",
                "name": "Muster, Anna",
                "jahrgang": "08",
                "verein": "Baden",
                "zeit": "00:01:08.88",
                "punkte": "852",
                "bemerkung": "Runde=A-Finale; Lauf=1; Bahn=4",
                "dq_status": "",
            },
            review_notes=[],
            raw_excerpt="1   Muster, Anna  Baden  08  1  4  1:08,88  852",
        )

        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            excel_path = ExcelExporter().export([record], temp_path / "records.xlsx", PROVISIONAL_EXPORT_FIELDS)
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            header = [cell.value for cell in worksheet[1]]
            first_row = [cell.value for cell in worksheet[2]]
            row = dict(zip(header, first_row, strict=True))

        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(row["name"], "Muster, Anna")
        self.assertEqual(row["gender"], "w")
        self.assertEqual(str(row["event_100m_obstacle_place"]), "1")
        self.assertEqual(row["event_100m_obstacle_time"], "1:08,88")
        self.assertEqual(str(row["event_100m_obstacle_points"]), "852")
        self.assertEqual(str(row["runde"]), "2")
        self.assertEqual(row["altersklasse"], "17/18")
        self.assertEqual(row["wertung"], "Einzelkampf")

    def test_summary_export_keeps_discipline_blocks_aligned_with_codes(self) -> None:
        record = ExtractedRecord(
            source_file=Path("bms.pdf"),
            page_number=2,
            parser_name="lifesaving",
            status=ExtractionStatus.PARTIAL,
            confidence=0.7,
            data={
                "wettkampfname": "BMS Frankenland 2025",
                "datum": "2025-03-14",
                "ort": "Koenigheim",
                "disziplin": "Einzelwertung",
                "altersklasse": "AK 10",
                "geschlecht": "maennlich",
                "platz": "5",
                "name": "Fleuchaus, Kilian",
                "jahrgang": "15",
                "verein": "Tauberbischofsheim",
                "punkte": "1028,95",
                "bemerkung": "Diff=1201; Disziplinen=50m k. Schwimmen | 50m Freistil | 50m Hindernis",
                "dq_status": "DQ",
                "summary_discipline_keys": "event_50m_combined_swim|event_50m_freestyle|event_50m_obstacle",
                "summary_discipline_count": "3",
            },
            review_notes=["Status aus Ergebniszeile abgeleitet: DQ"],
            raw_excerpt=(
                "5  Fleuchaus, Kilian  Tauberbischofsheim  15  1028,95  1201  "
                "0:32,01  S1: disq.  0:34,21  419  0:39,46  610"
            ),
        )

        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            excel_path = ExcelExporter().export([record], temp_path / "records.xlsx", PROVISIONAL_EXPORT_FIELDS)
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            header = [cell.value for cell in worksheet[1]]
            first_row = [cell.value for cell in worksheet[2]]
            row = dict(zip(header, first_row, strict=True))

        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(row["name"], "Fleuchaus, Kilian")
        self.assertEqual(row["altersklasse"], "10")
        self.assertEqual(row["wertung"], "3-Kampf")
        self.assertEqual(row["event_50m_combined_swim_time"], "0:32,01")
        self.assertEqual(row["event_50m_combined_swim_code"], "S1")
        self.assertEqual(row["event_50m_combined_swim_penalty"], "disq.")
        self.assertEqual(row["event_50m_freestyle_time"], "0:34,21")
        self.assertEqual(str(row["event_50m_freestyle_points"]), "419")
        self.assertEqual(row["event_50m_obstacle_time"], "0:39,46")
        self.assertEqual(str(row["event_50m_obstacle_points"]), "610")

    def test_summary_export_preserves_empty_choice_slots_in_higher_age_groups(self) -> None:
        record = ExtractedRecord(
            source_file=Path("bms.pdf"),
            page_number=8,
            parser_name="lifesaving",
            status=ExtractionStatus.PARTIAL,
            confidence=0.7,
            data={
                "wettkampfname": "BMS Frankenland 2025",
                "datum": "2025-03-14",
                "ort": "Wallduern",
                "disziplin": "Einzelwertung",
                "altersklasse": "AK 15/16",
                "geschlecht": "m",
                "platz": "1",
                "name": "Helmecke, Erik",
                "jahrgang": "09",
                "verein": "Adelsheim",
                "punkte": "1373,28",
                "bemerkung": "Diff=0; Disziplinen=100m Hindernis | 50m Retten | 100m Retten mit Flossen | 100m Lifesaver | 100m Kombi | 200m Super-Lifesaver",
                "dq_status": "DQ",
                "summary_discipline_keys": (
                    "event_100m_obstacle|event_50m_manikin_carry|event_100m_manikin_carry_with_fins|"
                    "event_100m_lifesaver|event_100m_rescue_medley|event_200m_super_lifesaver"
                ),
                "summary_discipline_count": "6",
            },
            review_notes=["Status aus Ergebniszeile abgeleitet: DQ"],
            raw_excerpt=(
                "1 Helmecke, Erik  Adelsheim  09  1373,28  0 "
                "... 1:31,20  513 ...  ... 1:54,03  S3 ... 1:38,23  400 ...  ... 3:27,05  461"
            ),
        )

        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            excel_path = ExcelExporter().export([record], temp_path / "records.xlsx", PROVISIONAL_EXPORT_FIELDS)
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            header = [cell.value for cell in worksheet[1]]
            first_row = [cell.value for cell in worksheet[2]]
            row = dict(zip(header, first_row, strict=True))

        self.assertEqual(row["wertung"], "4-Kampf")
        self.assertEqual(row["event_100m_obstacle_time"], "")
        self.assertEqual(row["event_50m_manikin_carry_time"], "1:31,20")
        self.assertEqual(str(row["event_50m_manikin_carry_points"]), "513")
        self.assertEqual(row["event_100m_manikin_carry_with_fins_time"], "")
        self.assertEqual(row["event_100m_lifesaver_time"], "1:54,03")
        self.assertEqual(row["event_100m_lifesaver_code"], "S3")
        self.assertEqual(row["event_100m_rescue_medley_time"], "1:38,23")
        self.assertEqual(str(row["event_100m_rescue_medley_points"]), "400")
        self.assertEqual(row["event_200m_super_lifesaver_time"], "3:27,05")
        self.assertEqual(str(row["event_200m_super_lifesaver_points"]), "461")

    def test_summary_export_handles_zero_points_before_na_status(self) -> None:
        record = ExtractedRecord(
            source_file=Path("bms.pdf"),
            page_number=8,
            parser_name="lifesaving",
            status=ExtractionStatus.PARTIAL,
            confidence=0.7,
            data={
                "wettkampfname": "BMS Frankenland 2025",
                "datum": "2025-03-14",
                "ort": "Wallduern",
                "disziplin": "Einzelwertung",
                "altersklasse": "AK 15/16",
                "geschlecht": "m",
                "platz": "2",
                "name": "Thies, Samuel",
                "jahrgang": "09",
                "verein": "Urphar",
                "punkte": "926,41",
                "bemerkung": "Diff=447",
                "dq_status": "DNS",
                "summary_discipline_keys": (
                    "event_100m_obstacle|event_50m_manikin_carry|event_100m_manikin_carry_with_fins|"
                    "event_100m_lifesaver|event_100m_rescue_medley|event_200m_super_lifesaver"
                ),
                "summary_discipline_count": "6",
            },
            review_notes=["Status aus Ergebniszeile abgeleitet: DNS"],
            raw_excerpt=(
                "2 Thies, Samuel  Urphar  09  926,41  447 "
                "...  ... 0:50,05  V1 ... 0  n.a. ... 1:23,74  479 ...  ... 3:31,65  448"
            ),
        )

        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            excel_path = ExcelExporter().export([record], temp_path / "records.xlsx", PROVISIONAL_EXPORT_FIELDS)
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            header = [cell.value for cell in worksheet[1]]
            first_row = [cell.value for cell in worksheet[2]]
            row = dict(zip(header, first_row, strict=True))

        self.assertEqual(row["wertung"], "4-Kampf")
        self.assertEqual(row["event_100m_obstacle_time"], "")
        self.assertEqual(row["event_50m_manikin_carry_time"], "")
        self.assertEqual(row["event_100m_manikin_carry_with_fins_time"], "0:50,05")
        self.assertEqual(row["event_100m_manikin_carry_with_fins_code"], "V1")
        self.assertEqual(str(row["event_100m_lifesaver_points"]), "0")
        self.assertEqual(row["event_100m_lifesaver_code"], "n.a.")
        self.assertEqual(row["event_100m_rescue_medley_time"], "1:23,74")
        self.assertEqual(str(row["event_100m_rescue_medley_points"]), "479")
        self.assertEqual(row["event_200m_super_lifesaver_time"], "3:31,65")
        self.assertEqual(str(row["event_200m_super_lifesaver_points"]), "448")


if __name__ == "__main__":
    unittest.main()
